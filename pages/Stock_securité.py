import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
import calendar
from datetime import datetime

# Charger le fichier Excel
chemin_fichier = "H:\\Drive partagés\\ACHAT 2024\\Streamlit_app\\Stock de sécurité.xlsx"
sm=pd.read_excel(chemin_fichier)

# Charger le fichier Excel
aujourd_hui = datetime.now()
Mois = aujourd_hui.month
Year = aujourd_hui.year
mois_en_lettres = calendar.month_name[Mois]

# Construisez le chemin du fichier Excel
excel_file_path = f"H:\\Drive partagés\\ACHAT 2024\\SITUATION STOCK\\Suivi situation stock {mois_en_lettres} {Year}.xlsx"

df=pd.read_excel(excel_file_path,sheet_name='situation')
df=df[['Description','St_Final']]

# Fusionner les deux Dataframe
wr=pd.merge(df,sm,on='Description', how='left')

# Sélectionnez les lignes où St_Final est inférieur ou égal à Alert de stock
selected_rows = wr[wr['St_Final'] <= wr['Alert de stock']]
selected_rows=selected_rows[['Description','St_Final']]
st.table(selected_rows) 

# Colorer les articles dont la quantité est inferieure au stock minimaum
wb = openpyxl.load_workbook(excel_file_path)
ws = wb['situation']

for index, row in wr.iterrows():
    cell_address = f'I{index+2}'
    if row['St_Final']<=row['Alert de stock']:
        fill_pattern=PatternFill(patternType='solid',fgColor='C64747')
        ws[cell_address].fill=fill_pattern
    else:
        # Supprimer le motif de remplissage en affectant None à l'attribut fill
        ws[cell_address].fill = PatternFill(fill_type='none')

# Sauvegarder le fichier modifié
wb.save(excel_file_path)