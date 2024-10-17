import pandas as pd
import streamlit as st
import numpy as np
import openpyxl
import re
import calendar
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle




st.title("Migrer les mouvement de stock vers la situation de stock mensuelle")

#===========================================================================================================================================
# Afficher un bouton de téléchargement de fichier Excel
btn_telecharger = st.file_uploader("Télécharger un fichier Excel", type=["xls"])

# Si le fichier est téléchargé
if btn_telecharger is not None:
    try:
        # Lire le fichier Excel
        df = pd.read_excel(btn_telecharger)

    except pd.errors.ParserError as e:
        st.error(f"Erreur lors de la lecture du fichier Excel : {e}")
        st.stop()  # Stop further execution
    
    # Specifier les colones a utiliser
    df=df[['Description', 'Quantité', 'Emplacement source', 'Emplacement de destination', 'Date']]

    # Fonction de nettoyage spécifique
    def clean_description(description):
        return re.sub(r'\[[^\]]*\]|\([^)]*\)', '', str(description)).strip()
    # Appliquer la fonction à la colonne 'Description'
    df['Description'] = df['Description'].apply(clean_description)

    # Charger les données à partir du fichier Excel pour le tableau de designation
    fichier_excel_designation = pd.ExcelFile("H:\\Drive partagés\\ACHAT 2024\\Streamlit_app\\liste_des_articles.xlsx")
    noms_feuilles_designation = fichier_excel_designation.sheet_names
    dfs_designation = [fichier_excel_designation.parse(nom_feuille) for nom_feuille in noms_feuilles_designation]
    df_designation = pd.concat(dfs_designation, ignore_index=True)

    #Fusionner avec le DataFrame des familles
    df_final = pd.merge(df, df_designation, on='Description', how='right')
    df_final=df_final.dropna(subset=['Quantité'])

#=========================================== Creer Nature de Mouvement======================================================
    data = {
        'Emplacement source': ["Stock Entrepôt principal [Entrepôt principal]", "Stock Production et packaging [Entrepôt principal]",
                               "Stock Process [Entrepôt principal]","STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]",
                               "Stock [Piece de rechange]","Inventory loss" ],
        'Emplacement de destination': ["Stock Entrepôt principal [Entrepôt principal]", "Stock Production et packaging [Entrepôt principal]",
                               "Stock Process [Entrepôt principal]","STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]",
                               "Stock [Piece de rechange]","Inventory loss" ]  
        }

    df_final['Nature de Mouvement'] = 'Supprimer' 
    conditions = [
        (df_final['Emplacement source'] == "Stock Process [Entrepôt principal]") & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]") & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Stock Process [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]"),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Suppliers') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Suppliers') & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Inventory loss') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Inventory loss') & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Suppliers'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Suppliers'),
        (df_final['Emplacement source'] == 'Stock Production et packaging [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Process [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]") & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Inventory loss'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Production et packaging [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Production et packaging [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Process [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Production et packaging [Entrepôt principal]') & (df_final['Emplacement de destination'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]"),
        (df_final['Emplacement source'] == 'Stock Production et packaging [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Stock Production et packaging [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Inventory loss'),
        (df_final['Emplacement source'] == 'Stock Process [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Process [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Process [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Process [Entrepôt principal]') & (df_final['Emplacement de destination'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]"),
        (df_final['Emplacement source'] == 'Stock Process [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Inventory loss'),
        (df_final['Emplacement source'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]") & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]'),
        (df_final['Emplacement source'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]") & (df_final['Emplacement de destination'] == 'Stock Process [Entrepôt principal]'),
        (df_final['Emplacement source'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]") & (df_final['Emplacement de destination'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]"),
        (df_final['Emplacement source'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]") & (df_final['Emplacement de destination'] == 'Inventory loss'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Inventory loss') & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Inventory loss') & (df_final['Emplacement de destination'] == 'Stock Process [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Inventory loss') & (df_final['Emplacement de destination'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]"),
        (df_final['Emplacement source'] == 'Inventory loss') & (df_final['Emplacement de destination'] == 'Inventory loss'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Stock Process [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == "STOCK SERVICE TRAITEMENT D'EAU [Entrepôt principal]"),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Inventory loss'),
        (df_final['Emplacement source'] == 'Stock Entrepôt principal [Entrepôt principal]') & (df_final['Emplacement de destination'] == 'Your Company: Transit Location'),
        (df_final['Emplacement source'] == 'Your Company: Transit Location') & (df_final['Emplacement de destination'] == 'Stock Entrepôt principal [Entrepôt principal]'),
        (df_final['Emplacement source'] == 'Stock [Piece de rechange]') & (df_final['Emplacement de destination'] == 'Your Company: Transit Location'),
        (df_final['Emplacement source'] == 'Your Company: Transit Location') & (df_final['Emplacement de destination'] == 'Stock [Piece de rechange]'),
        (df_final['Emplacement source'] == 'Production') & (df_final['Emplacement de destination'] == 'Stock Production et packaging [Entrepôt principal]')
        ]

    values = ['Erreure','Erreure','Erreure','Erreure','Erreure','Impossible','Impossible','Réception','Réception','Réception','Réception','Retour Fournisseur','Retour Fournisseur',
            'Retour Production','Retour Production','Retour Production','Transfer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer',
            'Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Supprimer','Transfer','Transfer','Transfer','Transfer','Transit location',
            'Transit location','Transit location','Transit location','Transfer'
            ]
    for condition, value in zip(conditions, values):
        df_final.loc[condition, 'Nature de Mouvement'] = value

    df_final = df_final[df_final['Nature de Mouvement'] != 'Supprimer']

    # Condition pour 'Jus coktail Kiddy 125 ml'
    condition1 = df_final['Description'] == 'Jus coktail Kiddy 125 ml'
    df_final.loc[condition1, 'Quantité'] = df_final.loc[condition1, 'Quantité'] / 156
    # Condition pour 'Jus Coktail Yummy 200 ml'
    condition2 = df_final['Description'] == 'Jus Coktail Yummy 200 ml'
    df_final.loc[condition2, 'Quantité'] = df_final.loc[condition2, 'Quantité'] / 150
    # Condition pour 'Jus Coktail Yummy 200 ml'    
    condition3 = df_final['Description'] == 'jus TOP YUMMY mangue 200ml'
    df_final.loc[condition3, 'Quantité'] = df_final.loc[condition3, 'Quantité'] / 150
    # Arrondire les chiffres
    df_final['Quantité'] = df_final['Quantité'].round(2)


#============================================= Etablire le pivot table ===========================================================

    pivot_table=pd.pivot_table(df_final, columns=['Nature de Mouvement'],values=['Quantité'], index=['Description'], aggfunc='sum', fill_value=0)
    pivot_table.columns = pivot_table.columns.get_level_values(1)
    pivot_table.reset_index(inplace=True)

    # voir si les columns sont existent
    columns_to_add = ['Retour Fournisseur', 'Retour Production','Réception','Transfer']

    for column in columns_to_add:
        if column not in pivot_table.columns:
            pivot_table[column] = 0
    pivot_table=pivot_table.loc[:,['Description','Retour Production','Réception','Transfer','Retour Fournisseur']]
    elements=np.array(pivot_table)

    #======================================== Extraire le nom de la feuille  ===========================================

    # Assurez que la colonne 'Date' est au format DateTime avec le format spécifique
    df_final['Date'] = pd.to_datetime(df_final['Date'], format='%d/%m/%Y %H:%M:%S', errors='coerce')

    # Extraire le jour dans une nouvelle colonne 'Jour'
    Jour = df_final['Date'].dt.day.unique()
    Mois = df_final['Date'].dt.month.unique()
    mois_en_lettres = [calendar.month_name[m] for m in Mois]
    Year = df_final['Date'].dt.year.unique()

    valeur_unique=Jour[0]
    Mois_unique=mois_en_lettres[0]
    Year_unique=Year[0]
    
#================================ Introduire les données dans le fihier suivis stock  ===========================================
    
    # Construisez le chemin du fichier Excel
    chemin_fichier_excel = f"H:\\Drive partagés\\ACHAT 2024\\SITUATION STOCK\\Suivi situation stock {Mois_unique} {Year_unique}.xlsx"

    # Ouvrir le classeur Excel
    wb = openpyxl.load_workbook(chemin_fichier_excel)

    # Obtenir la feuille de calcul active
    ws = wb[str(valeur_unique)]

    # Parcourir la liste d'éléments
    for element in elements:
        Description = element[0]
        Retour_Production=element[1]
        Réception = element[2]
        Transfer = element[3]
        Retour_Fournisseur = element[4]

        # Rechercher la description dans la feuille de calcul
        for row in ws.iter_rows():
            if row[1].value == Description:
                # Affecter les valeurs aux cellules Réception et Transfer
                ws.cell(row[0].row, 4).value = Retour_Production
                ws.cell(row[0].row, 5).value = Réception
                ws.cell(row[0].row, 6).value = Transfer
                ws.cell(row[0].row, 7).value = Retour_Fournisseur
                break  # Sortir de la boucle une fois que l'élément est trouvé
    
    # Afficher le fichier Excel
    st.write(pivot_table) 
    
    # Enregistrer le classeur Excel
    wb.save(chemin_fichier_excel)
    
else:
    st.warning("Veuillez télécharger un fichier Excel valide.")
  

# Fonction pour imprimer le tableau en format PDF
def print_to_pdf(data):
    st.markdown("**Impression en PDF en cours...**")

    # Convertir le DataFrame en une liste de listes
    table_data = [list(data.columns)] + data.values.tolist()

    # Créer un document PDF
    pdf_filename = "exported_table.pdf"
    pdf_buffer = BytesIO()
    pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)

    # Créer la table et définir le style
    table = Table(table_data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # En-tête de colonne en gris
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texte de l'en-tête de colonne en blanc
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alignement au centre
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Police en gras pour l'en-tête de colonne
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Ajouter un espace en bas de l'en-tête de colonne
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Couleur de fond des lignes impaires
        ])
    table.setStyle(style)

    # Ajouter la table au document PDF
    pdf.build([table])

    # Récupérer les données du PDF
    pdf_data = pdf_buffer.getvalue()

    # Télécharger le fichier PDF généré
    st.download_button(
        label="Télécharger PDF",
        data=pdf_data,
        key='pdf_export',
        file_name=pdf_filename,
        mime='application/pdf'
        )

# Bouton pour imprimer le tableau en format PDF
if st.button("Imprimer en PDF"):
    print_to_pdf(pivot_table)